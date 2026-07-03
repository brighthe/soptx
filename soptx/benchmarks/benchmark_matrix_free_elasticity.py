"""Benchmark SOPTX matrix-free elasticity across backends/devices.

This script measures the SOPTX matrix-free elasticity operator on multiple
fealpy backends/devices (NumPy CPU, PyTorch CPU, PyTorch CUDA) and records, per
configuration: assembled-vs-matrix-free matvec correctness, matvec timing, CG
solve timing/convergence, and a matrix-free memory estimate.

Design: each (backend, device) configuration runs in its own subprocess (worker
mode) so global backend state and the CUDA default-device setting never leak
across configurations. The driver process aggregates the per-config JSON rows
into a combined CSV and a formatted XLSX, including a speedup column relative to
the NumPy CPU baseline.

Conclusion (2026-06-30, RTX 5070 Ti / sm_120, torch 2.11.0+cu128)
-----------------------------------------------------------------
The SOPTX matrix-free elasticity path runs on all three configs (NumPy CPU /
PyTorch CPU / PyTorch CUDA) and produces consistent results across them: for
every case the matrix-free MatVec matches the assembled operator to rel_err
~1e-13..1e-15, and the converged small-scale displacement norm |u| agrees
bit-for-bit across the three configs.

On GPU the matrix-free MatVec shows a speedup trend that grows with problem
size (2D, vs NumPy CPU baseline):

    ndof      PyTorch CPU   PyTorch CUDA
    8,450        1.46x          2.16x
    33,282       2.35x          7.83x
    132,098      2.31x         11.91x   (~12x at ndof~1.3e5, still climbing)

Scope / caveats (do NOT overstate):
- The GPU speedup above is for the matrix-free MatVec operator only, not for
  end-to-end solve performance.
- The CG solver here is unpreconditioned; at large scale it does not converge
  within maxiter, so large-scale |u| is a truncated iterate (small cross-config
  float-order differences are expected and are not a correctness issue). A
  matrix-free preconditioner is the next prerequisite for end-to-end GPU solves.
- At small sizes the GPU is slower than NumPy (kernel-launch + Python CG-loop
  overhead dominate); the crossover starts around ndof ~ 1e4.

Examples
--------
# Run all three configs (numpy cpu / pytorch cpu / pytorch cuda) and aggregate:
python -m soptx.benchmarks.benchmark_matrix_free_elasticity

# Only NumPy + PyTorch CPU:
python -m soptx.benchmarks.benchmark_matrix_free_elasticity --configs numpy-cpu,pytorch-cpu
"""

from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from time import perf_counter
from typing import Callable, Optional

import numpy as np
from fealpy.backend import backend_manager as bm


@dataclass
class BenchmarkRow:
    case: str
    dim: int
    backend: str
    device: str
    ncell: int
    ndof: int
    nnz: int
    assembly_time_s: float
    assembled_matvec_time_s: float
    matrix_free_matvec_time_s: float
    rel_matvec_error: float
    cg_converged: bool
    cg_iterations: int
    cg_final_residual: float
    cg_rel_residual: float
    matrix_free_solve_time_s: float
    solution_norm: float
    assembled_memory_mb: float
    matrix_free_memory_est_mb: float


# --------------------------------------------------------------------------- #
# backend / device configuration (worker side)
# --------------------------------------------------------------------------- #

def configure_backend(backend: str, device: str) -> Callable[[], None]:
    """Set the fealpy backend/device for this process; return a sync callable."""
    if backend == "numpy":
        bm.set_backend("numpy")
        return lambda: None

    if backend == "pytorch":
        import torch
        torch.set_default_dtype(torch.float64)
        bm.set_backend("pytorch")
        if device == "cuda":
            if not torch.cuda.is_available():
                raise RuntimeError("CUDA requested but torch.cuda.is_available() is False")
            bm.set_default_device("cuda")
            return torch.cuda.synchronize
        bm.set_default_device("cpu")
        return lambda: None

    raise ValueError(f"Unsupported backend={backend!r}")


def _to_numpy(x):
    if hasattr(bm, "to_numpy"):
        return bm.to_numpy(x)
    return np.asarray(x)


def _time_average(fn: Callable[[], object], repeat: int,
                  sync: Callable[[], None]) -> tuple[object, float]:
    result = fn()        # warm-up (kernel build / allocation, esp. on CUDA)
    sync()
    start = perf_counter()
    for _ in range(repeat):
        result = fn()
    sync()
    elapsed = (perf_counter() - start) / repeat
    return result, elapsed


# --------------------------------------------------------------------------- #
# case construction
# --------------------------------------------------------------------------- #

def _build_case(dim: int, n: int):
    from fealpy.functionspace import LagrangeFESpace, TensorFunctionSpace
    from fealpy.mesh import TetrahedronMesh, TriangleMesh
    from soptx.analysis.integrators.linear_elastic_integrator import LinearElasticIntegrator
    from soptx.interpolation.linear_elastic_material import IsotropicLinearElasticMaterial

    if dim == 2:
        mesh = TriangleMesh.from_box(box=[0.0, 1.0, 0.0, 1.0], nx=n, ny=n)
        plane_type = "plane_stress"
    elif dim == 3:
        mesh = TetrahedronMesh.from_box(
            box=[0.0, 1.0, 0.0, 1.0, 0.0, 1.0], nx=n, ny=n, nz=n)
        plane_type = "3d"
    else:
        raise ValueError(f"Unsupported dim={dim}")

    fixed_threshold = lambda p: bm.abs(p[..., 0]) < 1.0e-12

    scalar_space = LagrangeFESpace(mesh, p=1, ctype="C")
    tensor_space = TensorFunctionSpace(scalar_space, shape=(-1, mesh.geo_dimension()))
    material = IsotropicLinearElasticMaterial(
        youngs_modulus=1.0, poisson_ratio=0.3, plane_type=plane_type)
    integrator = LinearElasticIntegrator(material=material, q=4, method="standard")
    coef = bm.linspace(0.4, 1.0, mesh.number_of_cells())
    integrator.coef = coef
    return mesh, tensor_space, integrator, coef, fixed_threshold


def _assembled_nnz_and_memory(K) -> tuple[int, float]:
    """nnz and an assembled-CSR-style memory estimate, device-agnostic."""
    values = K.values if not callable(getattr(K, "values", None)) else K.values()
    nnz = int(_to_numpy(values).shape[-1])
    # data (float64, 8B) + row/col indices (int64, 8B each)
    bytes_used = nnz * (8 + 8 + 8)
    return nnz, bytes_used / 1024**2


def _matrix_free_memory_est_mb(space, coef) -> float:
    cell2dof = space.cell_to_dof()
    bytes_used = _to_numpy(cell2dof).nbytes
    if coef is not None:
        bytes_used += _to_numpy(coef).nbytes
    return bytes_used / 1024**2


# --------------------------------------------------------------------------- #
# single case (worker side)
# --------------------------------------------------------------------------- #

def run_case(dim: int, n: int, backend: str, device: str, repeat: int,
             cg_tol: float, cg_maxiter: int, sync: Callable[[], None]) -> BenchmarkRow:
    from fealpy.fem import BilinearForm
    from soptx.analysis.matrix_free import MatrixFreeCGSolver, MatrixFreeElasticityOperator

    mesh, tensor_space, integrator, coef, fixed_threshold = _build_case(dim, n)
    case = f"{dim}d_n{n}"

    bform = BilinearForm(tensor_space)
    bform.add_integrator(integrator)

    sync()
    assembly_start = perf_counter()
    K = bform.assembly(format="coo")
    sync()
    assembly_time = perf_counter() - assembly_start

    nnz, assembled_memory_mb = _assembled_nnz_and_memory(K)
    mf_memory_mb = _matrix_free_memory_est_mb(tensor_space, coef)

    # from here on the matrix-free action must use contraction, never local Ke@xe
    integrator._disable_action_assembly_fallback = True

    op = MatrixFreeElasticityOperator(space=tensor_space, integrator=integrator, rho=coef)
    ndof = tensor_space.number_of_global_dofs()
    x = bm.linspace(0.1, 1.0, ndof)

    y_ref, assembled_matvec_time = _time_average(lambda: K.matmul(x), repeat, sync)
    y_mf, mf_matvec_time = _time_average(lambda: op.matvec(x), repeat, sync)
    rel_matvec_error = float(bm.linalg.norm(y_ref - y_mf) / bm.linalg.norm(y_ref))

    fixed_dofs = bm.where(
        tensor_space.is_boundary_dof(threshold=fixed_threshold, method="interp"))[0]
    rhs = bm.linspace(0.2, 1.1, ndof)
    rhs = bm.set_at(rhs, fixed_dofs, 0.0)

    op_bc = MatrixFreeElasticityOperator(
        space=tensor_space, integrator=integrator, rho=coef, dirichlet_dofs=fixed_dofs)
    solver = MatrixFreeCGSolver(tol=cg_tol, maxiter=cg_maxiter)

    sync()
    solve_start = perf_counter()
    u_mf, info = solver.solve(op_bc, rhs)
    sync()
    solve_time = perf_counter() - solve_start

    rel_residual = info.final_residual / max(info.initial_residual, 1e-300)
    solution_norm = float(bm.linalg.norm(u_mf))

    return BenchmarkRow(
        case=case, dim=dim, backend=backend, device=device,
        ncell=mesh.number_of_cells(), ndof=ndof, nnz=nnz,
        assembly_time_s=assembly_time,
        assembled_matvec_time_s=assembled_matvec_time,
        matrix_free_matvec_time_s=mf_matvec_time,
        rel_matvec_error=rel_matvec_error,
        cg_converged=info.converged, cg_iterations=info.iterations,
        cg_final_residual=info.final_residual, cg_rel_residual=rel_residual,
        matrix_free_solve_time_s=solve_time, solution_norm=solution_norm,
        assembled_memory_mb=assembled_memory_mb,
        matrix_free_memory_est_mb=mf_memory_mb,
    )


def _parse_sizes(text: str) -> list[int]:
    if not text:
        return []
    return [int(part.strip()) for part in text.split(",") if part.strip()]


# --------------------------------------------------------------------------- #
# worker entry: run one (backend, device), emit JSON rows on stdout
# --------------------------------------------------------------------------- #

def run_worker(args: argparse.Namespace) -> None:
    sync = configure_backend(args.backend, args.device)
    for n in _parse_sizes(args.cases_2d):
        row = run_case(2, n, args.backend, args.device, args.repeat,
                       args.cg_tol, args.cg_maxiter, sync)
        print("ROW " + json.dumps(asdict(row)), flush=True)
    for n in _parse_sizes(args.cases_3d):
        row = run_case(3, n, args.backend, args.device, args.repeat,
                       args.cg_tol, args.cg_maxiter, sync)
        print("ROW " + json.dumps(asdict(row)), flush=True)


# --------------------------------------------------------------------------- #
# driver entry: spawn one worker per config, aggregate
# --------------------------------------------------------------------------- #

CONFIGS = {
    "numpy-cpu": ("numpy", "cpu"),
    "pytorch-cpu": ("pytorch", "cpu"),
    "pytorch-cuda": ("pytorch", "cuda"),
}


def _cuda_available() -> bool:
    try:
        import torch
        return torch.cuda.is_available()
    except Exception:
        return False


def _spawn_worker(backend: str, device: str, args: argparse.Namespace) -> list[dict]:
    cmd = [
        sys.executable, "-m", "soptx.benchmarks.benchmark_matrix_free_elasticity",
        "--worker", "--backend", backend, "--device", device,
        "--cases-2d", args.cases_2d, "--cases-3d", args.cases_3d,
        "--repeat", str(args.repeat),
        "--cg-tol", str(args.cg_tol), "--cg-maxiter", str(args.cg_maxiter),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        print(f"  [WARN] config {backend}-{device} failed (rc={proc.returncode}):",
              file=sys.stderr)
        tail = "\n".join(proc.stderr.strip().splitlines()[-8:])
        print(tail, file=sys.stderr)
        return []
    rows = []
    for line in proc.stdout.splitlines():
        if line.startswith("ROW "):
            rows.append(json.loads(line[4:]))
    return rows


def _add_speedup(rows: list[dict]) -> list[dict]:
    """Add matvec/solve speedup vs the numpy-cpu baseline for the same case."""
    baseline = {}
    for r in rows:
        if r["backend"] == "numpy" and r["device"] == "cpu":
            baseline[r["case"]] = r
    for r in rows:
        base = baseline.get(r["case"])
        if base and r["matrix_free_matvec_time_s"] > 0:
            r["matvec_speedup_vs_numpy"] = (
                base["matrix_free_matvec_time_s"] / r["matrix_free_matvec_time_s"])
            r["solve_speedup_vs_numpy"] = (
                base["matrix_free_solve_time_s"] / max(r["matrix_free_solve_time_s"], 1e-300))
        else:
            r["matvec_speedup_vs_numpy"] = float("nan")
            r["solve_speedup_vs_numpy"] = float("nan")
    return rows


def _write_csv(rows: list[dict], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(rows: list[dict], output: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output.parent.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "matrix_free_benchmark"
    ws.append(headers)
    for item in rows:
        ws.append([item[key] for key in headers])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    scientific_cols = {"rel_matvec_error", "cg_final_residual", "cg_rel_residual"}
    time_cols = {"assembly_time_s", "assembled_matvec_time_s",
                 "matrix_free_matvec_time_s", "matrix_free_solve_time_s"}
    memory_cols = {"assembled_memory_mb", "matrix_free_memory_est_mb"}
    float_cols = {"solution_norm", "matvec_speedup_vs_numpy", "solve_speedup_vs_numpy"}
    integer_cols = {"dim", "ncell", "ndof", "nnz", "cg_iterations"}

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            max_len = max(max_len, len(str(cell.value)))
            if header in scientific_cols:
                cell.number_format = "0.000E+00"
            elif header in time_cols:
                cell.number_format = "0.000000"
            elif header in memory_cols:
                cell.number_format = "0.000"
            elif header in float_cols:
                cell.number_format = "0.000"
            elif header in integer_cols:
                cell.number_format = "0"
            elif header == "cg_converged":
                cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 24)

    wb.save(output)


def run_driver(args: argparse.Namespace) -> None:
    requested = [c.strip() for c in args.configs.split(",") if c.strip()]
    for c in requested:
        if c not in CONFIGS:
            raise ValueError(f"Unknown config {c!r}; choose from {list(CONFIGS)}")

    rows: list[dict] = []
    for c in requested:
        backend, device = CONFIGS[c]
        if device == "cuda" and not _cuda_available():
            print(f"  [SKIP] {c}: CUDA not available in this environment.")
            continue
        print(f"  running config {c} ({backend}/{device}) ...")
        rows.extend(_spawn_worker(backend, device, args))

    if not rows:
        print("No benchmark rows produced.")
        return

    rows = _add_speedup(rows)

    csv_output = Path(args.output)
    xlsx_output = Path(args.xlsx_output) if args.xlsx_output else csv_output.with_suffix(".xlsx")
    _write_csv(rows, csv_output)
    _write_xlsx(rows, xlsx_output)

    print(f"Wrote {len(rows)} rows to {csv_output}")
    print(f"Wrote formatted workbook to {xlsx_output}")
    for r in rows:
        print(
            f"{r['backend']}/{r['device']} {r['case']}: ndof={r['ndof']}, "
            f"mf_matvec={r['matrix_free_matvec_time_s']:.3e}s "
            f"(x{r['matvec_speedup_vs_numpy']:.2f} vs numpy), "
            f"rel_matvec={r['rel_matvec_error']:.2e}, "
            f"cg={r['cg_iterations']} (conv={r['cg_converged']}), "
            f"|u|={r['solution_norm']:.6e}"
        )


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run SOPTX matrix-free elasticity multi-backend benchmark.")
    parser.add_argument("--worker", action="store_true",
                        help="Internal: run a single (backend, device) and emit JSON rows.")
    parser.add_argument("--backend", default="numpy", choices=["numpy", "pytorch"])
    parser.add_argument("--device", default="cpu", choices=["cpu", "cuda"])
    parser.add_argument("--configs", default="numpy-cpu,pytorch-cpu,pytorch-cuda",
                        help="Comma-separated configs for driver mode: "
                             "numpy-cpu,pytorch-cpu,pytorch-cuda")
    parser.add_argument("--output", default="outputs/matrix_free_elasticity_benchmark.csv")
    parser.add_argument("--xlsx-output", default=None)
    parser.add_argument("--repeat", type=int, default=5)
    parser.add_argument("--cases-2d", default="4,8,16,32")
    parser.add_argument("--cases-3d", default="2,4,8")
    parser.add_argument("--cg-tol", type=float, default=1.0e-10)
    parser.add_argument("--cg-maxiter", type=int, default=1000)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.worker:
        run_worker(args)
    else:
        run_driver(args)


if __name__ == "__main__":
    main()
